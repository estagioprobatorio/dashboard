import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def build_spreadsheet():
    wb = openpyxl.Workbook()
    
    # Estilos Visuais SEED-PR
    font_title = Font(name='Arial', size=14, bold=True, color='002D5C')
    font_subtitle = Font(name='Arial', size=11, italic=True, color='5A6E85')
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    font_bold = Font(name='Arial', size=10, bold=True)
    font_regular = Font(name='Arial', size=10)
    
    fill_navy = PatternFill(start_color='002D5C', end_color='002D5C', fill_type='solid')
    fill_blue_header = PatternFill(start_color='0B3C5D', end_color='0B3C5D', fill_type='solid')
    fill_warning = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    fill_success = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    fill_alert = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    
    thin_border_side = Side(border_style="thin", color="D1D5DB")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    # ----------------------------------------------------
    # ABA 1: LEIA-ME
    # ----------------------------------------------------
    ws_readme = wb.active
    ws_readme.title = "LEIA-ME"
    ws_readme.views.sheetView[0].showGridLines = True
    
    ws_readme['A1'] = "GESTÃO DA VIDA FUNCIONAL - ESTÁGIO PROBATÓRIO (SEED-PR)"
    ws_readme['A1'].font = font_title
    
    ws_readme['A2'] = "Manual de Instruções e Regras de Cálculo da Planilha Administrativa"
    ws_readme['A2'].font = font_subtitle
    
    instructions = [
        ("REGRA DOS 36 MESES", "O Estágio Probatório possui duração regulamentar de 36 meses (3 anos / 1095 dias)."),
        ("DIAS DE AFASTAMENTO", "Licenças ou afastamentos legais inseridos na coluna 'dias_afastamento' prorrogam a data final do estágio automaticamente na mesma proporção."),
        ("REGRA DA AVALIAÇÃO FORMATIVA (-90 DIAS)", "A Avaliação Formativa deve ser totalmente concluída EXATAMENTE 90 DIAS ANTES da data de término do Estágio Probatório."),
        ("FASE FINAL DE AVALIAÇÃO", "Quando o cursista atinge os últimos 90 dias do estágio probatório, seu status muda automaticamente para 'FASE FINAL (Últimos 90 dias)', sinalizando à equipe pedagógica que a avaliação formativa final precisa ser encerrada."),
        ("NÍVEL / ANO DO CURSISTA", "Calculado dinamicamente: Nível 1 (1º Ano), Nível 2 (2º Ano) ou Nível 3 (3º Ano) de acordo com o tempo decorrido do concurso/estágio."),
        ("APP / CARD DO CURSISTA", "Os dados desta planilha alimentam visualmente o App / Card Gamificado de Vida Funcional exibido para cada cursista no portal.")
    ]
    
    ws_readme.cell(row=4, column=1, value="TÓPICO").font = font_header
    ws_readme.cell(row=4, column=1).fill = fill_navy
    ws_readme.cell(row=4, column=2, value="DESCRIÇÃO E REGRA DE NEGÓCIO").font = font_header
    ws_readme.cell(row=4, column=2).fill = fill_navy
    
    for idx, (topic, desc) in enumerate(instructions, start=5):
        c1 = ws_readme.cell(row=idx, column=1, value=topic)
        c2 = ws_readme.cell(row=idx, column=2, value=desc)
        c1.font = font_bold
        c2.font = font_regular
        c1.border = border_cell
        c2.border = border_cell
        
    ws_readme.column_dimensions['A'].width = 38
    ws_readme.column_dimensions['B'].width = 90

    # ----------------------------------------------------
    # ABA 2: VIDA_FUNCIONAL_ADM
    # ----------------------------------------------------
    ws_main = wb.create_sheet(title="VIDA_FUNCIONAL_ADM")
    ws_main.views.sheetView[0].showGridLines = True
    
    headers = [
        "CGM", "Nome do Cursista", "CPF", "E-mail Google", "NRE",
        "Componente Curricular", "Instituição Lotação", "Instituição Exercício",
        "Vínculo", "Carga Horária", "ID Turma", "Formador Responsável",
        "Início Concurso", "Início Estágio", "Dias Afastamento", "Nível (Ano)",
        "Fim Estágio Previsto", "Data Limite Avaliação Formativa (-90d)",
        "Dias Restantes Estágio", "Dias Restantes Avaliação Formativa",
        "Status Avaliação Formativa", "Status Estágio Probatório"
    ]
    
    ws_main.row_dimensions[1].height = 32
    for col_num, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col_num, value=header)
        cell.font = font_header
        cell.fill = fill_blue_header
        cell.alignment = align_center
        cell.border = border_cell

    # Exemplo de Cursistas para Teste e Validação
    sample_cursistas = [
        (1029384, "Jorge Inacio Dotti", "123.456.789-00", "jorge.dotti@escola.pr.gov.br", "NRE Curitiba", "Matemática", "Colégio Estadual do Paraná", "Colégio Estadual do Paraná", "QPM", "20h", "TURMA-MAT-01", "Ana Maria Silva", "2024-07-12", "2024-07-12", 0),
        (2049581, "Mariana Castro de Oliveira", "987.654.321-11", "mariana.castro@escola.pr.gov.br", "NRE Londrina", "Língua Portuguesa", "C.E. Vicente Rijo", "C.E. Vicente Rijo", "QPM", "40h", "TURMA-POR-02", "Carlos Eduardo", "2023-09-01", "2023-09-01", 15),
        (3058492, "Roberto Carlos Ferreira", "456.789.012-22", "roberto.ferreira@escola.pr.gov.br", "NRE Maringá", "História", "C.E. Dr. Gastão Vidigal", "C.E. Dr. Gastão Vidigal", "QPM", "20h", "TURMA-HIS-01", "Fernanda Souza", "2022-02-15", "2022-02-15", 30),
        (4019283, "Beatriz Souza Mendes", "789.012.345-33", "beatriz.mendes@escola.pr.gov.br", "NRE Cascavel", "Biologia", "C.E. Wilson Joffre", "C.E. Wilson Joffre", "PSS", "20h", "TURMA-BIO-01", "Juliana Lima", "2025-01-20", "2025-01-20", 0)
    ]
    
    for row_idx, row_data in enumerate(sample_cursistas, start=2):
        ws_main.row_dimensions[row_idx].height = 22
        
        # Preenche dados estáticos
        for c_idx in range(15):
            val = row_data[c_idx]
            cell = ws_main.cell(row=row_idx, column=c_idx+1, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_idx in [0, 8, 9, 10, 14]:
                cell.alignment = align_center
            elif c_idx in [12, 13]:
                cell.alignment = align_center
                cell.number_format = 'yyyy-mm-dd'

        # Fórmulas Excel Dinâmicas
        # Col P (16): Nível (Ano): =MIN(3, MAX(1, INT(YEARFRAC(N2, TODAY(), 1))+1))
        f_nivel = f'=IF(N{row_idx}="","",MIN(3,MAX(1,INT(YEARFRAC(N{row_idx},TODAY(),1))+1)))'
        cell_p = ws_main.cell(row=row_idx, column=16, value=f_nivel)
        cell_p.font = font_bold
        cell_p.alignment = align_center
        cell_p.border = border_cell

        # Col Q (17): Fim Estágio Previsto: =EDATE(N2, 36) - 1 + O2
        f_fim_estagio = f'=IF(N{row_idx}="","",EDATE(N{row_idx},36)-1+O{row_idx})'
        cell_q = ws_main.cell(row=row_idx, column=17, value=f_fim_estagio)
        cell_q.font = font_bold
        cell_q.alignment = align_center
        cell_q.number_format = 'yyyy-mm-dd'
        cell_q.border = border_cell

        # Col R (18): Data Limite Avaliação Formativa (-90d): =Q2 - 90
        f_limite_af = f'=IF(Q{row_idx}="","",Q{row_idx}-90)'
        cell_r = ws_main.cell(row=row_idx, column=18, value=f_limite_af)
        cell_r.font = font_bold
        cell_r.alignment = align_center
        cell_r.number_format = 'yyyy-mm-dd'
        cell_r.border = border_cell

        # Col S (19): Dias Restantes Estágio: =MAX(0, Q2 - TODAY())
        f_dias_estagio = f'=IF(Q{row_idx}="","",MAX(0,Q{row_idx}-TODAY()))'
        cell_s = ws_main.cell(row=row_idx, column=19, value=f_dias_estagio)
        cell_s.font = font_bold
        cell_s.alignment = align_center
        cell_s.border = border_cell

        # Col T (20): Dias Restantes Avaliação Formativa: =MAX(0, R2 - TODAY())
        f_dias_af = f'=IF(R{row_idx}="","",MAX(0,R{row_idx}-TODAY()))'
        cell_t = ws_main.cell(row=row_idx, column=20, value=f_dias_af)
        cell_t.font = font_bold
        cell_t.alignment = align_center
        cell_t.border = border_cell

        # Col U (21): Status Avaliação Formativa
        f_status_af = f'=IF(R{row_idx}="","",IF(TODAY()>R{row_idx},"FINALIZADA (Prazo Excedido)",IF(T{row_idx}<=30,"ALERTA: Fim Próximo (30d)","EM ANDAMENTO")))'
        cell_u = ws_main.cell(row=row_idx, column=21, value=f_status_af)
        cell_u.font = font_bold
        cell_u.alignment = align_center
        cell_u.border = border_cell

        # Col V (22): Status Estágio Probatório
        f_status_ep = f'=IF(S{row_idx}="","",IF(S{row_idx}=0,"CONCLUÍDO",IF(TODAY()>=R{row_idx},"FASE FINAL (Últimos 90 dias)","EM ANDAMENTO")))'
        cell_v = ws_main.cell(row=row_idx, column=22, value=f_status_ep)
        cell_v.font = font_bold
        cell_v.alignment = align_center
        cell_v.border = border_cell

    # Ajusta largura de colunas automaticamente
    for col in ws_main.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_main.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # ----------------------------------------------------
    # ABA 3: DASHBOARD_INDICADORES
    # ----------------------------------------------------
    ws_dash = wb.create_sheet(title="DASHBOARD_INDICADORES")
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash['A1'] = "PAINEL DE INDICADORES DE GESTÃO DA VIDA FUNCIONAL"
    ws_dash['A1'].font = font_title
    
    dash_items = [
        ("TOTAL DE CURSISTAS CADASTRADOS", "=COUNTA(VIDA_FUNCIONAL_ADM!A2:A1000)"),
        ("CURSISTAS EM ANDAMENTO", '=COUNTIF(VIDA_FUNCIONAL_ADM!V2:V1000, "EM ANDAMENTO")'),
        ("CURSISTAS NA FASE FINAL DE AVALIAÇÃO (ÚLTIMOS 90 DIAS)", '=COUNTIF(VIDA_FUNCIONAL_ADM!V2:V1000, "FASE FINAL (Últimos 90 dias)")'),
        ("ESTÁGIOS PROBATÓRIOS CONCLUÍDOS", '=COUNTIF(VIDA_FUNCIONAL_ADM!V2:V1000, "CONCLUÍDO")'),
        ("AVALIAÇÕES FORMATIVAS COM PRAZO EXCEDIDO", '=COUNTIF(VIDA_FUNCIONAL_ADM!U2:U1000, "FINALIZADA (Prazo Excedido)")')
    ]
    
    ws_dash.cell(row=3, column=1, value="INDICADOR").font = font_header
    ws_dash.cell(row=3, column=1).fill = fill_navy
    ws_dash.cell(row=3, column=2, value="QUANTIDADE").font = font_header
    ws_dash.cell(row=3, column=2).fill = fill_navy
    
    for idx, (label, formula) in enumerate(dash_items, start=4):
        c1 = ws_dash.cell(row=idx, column=1, value=label)
        c2 = ws_dash.cell(row=idx, column=2, value=formula)
        c1.font = font_bold
        c2.font = font_bold
        c1.border = border_cell
        c2.border = border_cell
        c2.alignment = align_center
        
    ws_dash.column_dimensions['A'].width = 58
    ws_dash.column_dimensions['B'].width = 20
    
    # Salvar o arquivo
    out_dir = r"c:\Users\pseudocelomado\Documents\DESENVOLVIMENTO\ESTAGIO\dashboard-tawny-delta-96.vercel.app\dashboard-main\card"
    out_path = os.path.join(out_dir, "vida_funcional_estagio_probatorio_adm.xlsx")
    wb.save(out_path)
    print(f"Planilha criada com sucesso em: {out_path}")

if __name__ == "__main__":
    build_spreadsheet()
